import openpyxl as xl
import os
import copy
from openpyxl.styles import Font, colors, Alignment
from openpyxl.utils import get_column_letter

if __name__ == '__main__':
    file0 = "D://Sheet1.xlsx"
    file1 = "D://7246-督查督办系统20230309-复审1次.xlsx"
    if not os.path.exists(file0):
        print("数据源文件不存在")
    if not os.path.exists(file1):
        print("目标文件不存在")
    workbook0 = xl.load_workbook(file0)
    workbook1 = xl.load_workbook(file1)

    wb0_s1 = workbook0["Tables"]
    wb0_s2 = workbook0["Columns"]
    wb1_s1_bak = workbook1["M02元数据-系统数据表"]
    wb1_s2_bak = workbook1["M03元数据-数据字典"]
    wb1_s1 = workbook1["M02元数据-系统数据表"]
    wb1_s2 = workbook1["M03元数据-数据字典"]

    #save bak

    max_row = wb1_s1_bak.max_row
    max_column = wb1_s1_bak.max_column
    sheetnames=workbook1.sheetnames;
    for sheetname in sheetnames:
        sheet = workbook1[sheetname]
        max_row= sheet.max_row
        max_column= sheet.max_column
        for i in range(1, max_row):
            for j in range(1, max_column):
                wb1_s1.cell(i, j).value = copy.copy(wb0_s1.cell(i, j).value)
                wb1_s1.cell(i, j).style = copy.copy(wb0_s1.cell(i, j).style)



    # max_row = wb0_s1.max_row
    try:
    #     for i in range(1, max_row):
    #         cell = wb0_s1.cell(i + 1, 1)
    #         if cell.value is None:
    #             continue
    #         owner = "" + cell.value
    #         owner = owner.replace("User '", "").replace("'", "")
    #         wb1_s1.cell(i + 2, 2).value = owner
    #         wb1_s1.cell(i + 2, 2).style = copy.copy(wb0_s1.cell(i + 2, 2).style)
    #         wb1_s1.cell(i + 2, 2).font = copy.copy(wb0_s1.cell(i + 2, 2).font)
    #         wb1_s1.cell(i + 2, 2).border = copy.copy(wb0_s1.cell(i + 2, 2).border)
    #         wb1_s1.cell(i + 2, 2).fill = copy.copy(wb0_s1.cell(i + 2, 2).fill)
    #         wb1_s1.cell(i + 2, 2).number_format = copy.copy(wb0_s1.cell(i + 2, 2).number_format)
    #         wb1_s1.cell(i + 2, 2).protection = copy.copy(wb0_s1.cell(i + 2, 2).protection)
    #         wb1_s1.cell(i + 2, 2).alignment = copy.copy(wb0_s1.cell(i + 2, 2).alignment)
    #
    #         wb1_s1.cell(i+2, 4).value = wb0_s1.cell(i+1, 3).value
    #         wb1_s1.cell(i+2, 5).value = wb0_s1.cell(i+1, 2).value
    #         wb1_s1.cell(i+2, 6).value = wb0_s1.cell(i+1, 4).value
        workbook1.save(file1)
        workbook0.close()
        workbook1.close()
    except Exception as e:
        print(e)
