import openpyxl as xl
import os

if __name__ == '__main__':
    file0 = "D://7246-督查督办系统20230309-复审1次.xlsx"
    file1 = "D://test.xlsx"
    workbook0 = xl.load_workbook(file0)
    if os.path.exists(file1):
        workbook1 = xl.load_workbook(file1)
    else:
        workbook1 = xl.Workbook()
        workbook1.save(file1)

    wb0_s0 = workbook0["M01元数据-应用域和系统"]
    wb0_s1 = workbook0["M02元数据-系统数据表"]
    wb0_s2 = workbook0["M03元数据-数据字典"]

    sheetname = workbook1.sheetnames

    if "M01元数据-应用域和系统" not in sheetname:
        workbook1.create_sheet("M01元数据-应用域和系统", 0)
    if "M02元数据-系统数据表" not in sheetname:
        workbook1.create_sheet("M02元数据-系统数据表", 1)
    if "M03元数据-数据字典" not in sheetname:
        workbook1.create_sheet("M03元数据-数据字典", 2)
    if "Sheet1" in sheetname:
        workbook1.remove(workbook1["Sheet1"])

    wb1_s0 = workbook1.worksheets[0]
    wb1_s1 = workbook1.worksheets[1]
    wb1_s2 = workbook1.worksheets[2]
    max_row = wb0_s0.max_row
    max_column = wb0_s0.max_column
    for i in range(1, max_row):
        for j in range(1, max_column):
            wb1_s0.cell(i, j).value = wb0_s0.cell(i, j).value
            wb1_s0.cell(i, j).style = wb0_s0.cell(i, j).style

    max_row = wb0_s1.max_row
    max_column = wb0_s1.max_column
    for i in range(1, max_row):
        for j in range(1, max_column):
            wb1_s1.cell(i, j).value = wb0_s1.cell(i, j).value
            wb1_s1.cell(i, j).style = wb0_s1.cell(i, j).style

            # print(wb0_s0.cell(i, j).value)
    max_row = wb0_s2.max_row
    max_column = wb0_s2.max_column
    for i in range(1, max_row):
        for j in range(1, max_column):
            wb1_s2.cell(i, j).value = wb0_s2.cell(i, j).value
            # wb1_s2.cell(i, j).style = wb0_s2.cell(i, j).style
    workbook1.save("D://test.xlsx")
    workbook0.close()
    workbook1.close()
